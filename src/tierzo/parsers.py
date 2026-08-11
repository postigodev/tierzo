from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable, Literal

from openpyxl import load_workbook


SUPPORTED_INPUTS = {".csv", ".txt", ".xlsx"}


class InputLimitError(ValueError):
    def __init__(
        self,
        kind: Literal["too_many_items", "item_too_long"],
        limit: int,
        item_index: int | None = None,
    ) -> None:
        self.kind = kind
        self.limit = limit
        self.item_index = item_index
        super().__init__(kind)


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()


def collect_items(
    values: Iterable[object],
    *,
    flatten_whitespace: bool = False,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    items: list[str] = []
    for value in values:
        text = normalize_text(value)
        if flatten_whitespace:
            text = re.sub(r"\s+", " ", text).strip()
        if not text:
            continue

        item_index = len(items)
        if max_item_length is not None and len(text) > max_item_length:
            raise InputLimitError("item_too_long", max_item_length, item_index)
        if max_items is not None and item_index >= max_items:
            raise InputLimitError("too_many_items", max_items)
        items.append(text)
    return items


def parse_text_lines(
    raw_text: str,
    *,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    return collect_items(
        raw_text.splitlines(),
        max_items=max_items,
        max_item_length=max_item_length,
    )


def parse_txt_file(
    path: Path,
    *,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    with path.open("r", encoding="utf-8-sig") as file:
        return collect_items(
            file,
            max_items=max_items,
            max_item_length=max_item_length,
        )


def parse_csv_file(
    path: Path,
    *,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        return collect_items(
            (row[0] for row in reader if row),
            flatten_whitespace=True,
            max_items=max_items,
            max_item_length=max_item_length,
        )


def parse_xlsx_file(
    path: Path,
    *,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    workbook = load_workbook(
        path,
        data_only=True,
        keep_links=False,
        read_only=True,
    )
    try:
        sheet = workbook.worksheets[0]
        return collect_items(
            (
                row[0]
                for row in sheet.iter_rows(
                    min_col=1,
                    max_col=1,
                    values_only=True,
                )
            ),
            flatten_whitespace=True,
            max_items=max_items,
            max_item_length=max_item_length,
        )
    finally:
        workbook.close()


def parse_input_file(
    path: Path,
    *,
    max_items: int | None = None,
    max_item_length: int | None = None,
) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return parse_txt_file(
            path,
            max_items=max_items,
            max_item_length=max_item_length,
        )
    if suffix == ".csv":
        return parse_csv_file(
            path,
            max_items=max_items,
            max_item_length=max_item_length,
        )
    if suffix == ".xlsx":
        return parse_xlsx_file(
            path,
            max_items=max_items,
            max_item_length=max_item_length,
        )

    supported = ", ".join(sorted(SUPPORTED_INPUTS))
    raise ValueError(f"Unsupported input file type '{suffix}'. Supported types: {supported}")


def iter_non_empty(values: Iterable[str]) -> list[str]:
    return [value for value in (normalize_text(value) for value in values) if value]
