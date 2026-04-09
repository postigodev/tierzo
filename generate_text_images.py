from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


DEFAULT_SIZE = 1024
DEFAULT_PADDING_RATIO = 0.1
DEFAULT_FONT_SIZE = 180
MIN_FONT_SIZE = 20
LINE_SPACING_RATIO = 0.2


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read the first column of an .xlsx file and create one square PNG "
            "per non-empty cell with centered text."
        )
    )
    parser.add_argument("xlsx_file", type=Path, help="Path to the input .xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output folder for generated images. Defaults to <xlsx_stem>_images",
    )
    parser.add_argument(
        "--size",
        type=int,
        default=DEFAULT_SIZE,
        help=f"Square image size in pixels. Default: {DEFAULT_SIZE}",
    )
    parser.add_argument(
        "--background",
        default="#FFFFFF",
        help="Background color. Default: #FFFFFF",
    )
    parser.add_argument(
        "--text-color",
        default="#000000",
        help="Text color. Default: #000000",
    )
    parser.add_argument(
        "--font",
        type=Path,
        help="Optional path to a .ttf or .otf font file",
    )
    return parser.parse_args()


def get_font(font_path: Path | None, size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    if font_path:
        return ImageFont.truetype(str(font_path), size=size)

    common_fonts = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/segoeui.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
    ]

    for candidate in common_fonts:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)

    return ImageFont.load_default()


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()


def iter_first_column_values(xlsx_file: Path) -> Iterable[str]:
    workbook = load_workbook(xlsx_file, data_only=True)
    sheet = workbook.worksheets[0]

    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        text = normalize_text(row[0])
        if text:
            yield text


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    explicit_lines = text.split("\n")
    wrapped_lines: list[str] = []

    for line in explicit_lines:
        words = line.split()
        if not words:
            wrapped_lines.append("")
            continue

        current = words[0]
        for word in words[1:]:
            trial = f"{current} {word}"
            trial_width = draw.textbbox((0, 0), trial, font=font)[2]
            if trial_width <= max_width:
                current = trial
            else:
                wrapped_lines.append(current)
                current = word
        wrapped_lines.append(current)

    return wrapped_lines


def measure_text_block(
    draw: ImageDraw.ImageDraw,
    lines: list[str],
    font: ImageFont.ImageFont,
) -> tuple[int, int]:
    widths: list[int] = []
    heights: list[int] = []

    for line in lines:
        left, top, right, bottom = draw.textbbox((0, 0), line or " ", font=font)
        widths.append(right - left)
        heights.append(bottom - top)

    line_height = max(heights) if heights else 0
    line_spacing = max(4, int(line_height * LINE_SPACING_RATIO))
    total_height = sum(heights) + max(0, len(lines) - 1) * line_spacing
    total_width = max(widths) if widths else 0
    return total_width, total_height


def fit_text_to_box(
    image_size: int,
    text: str,
    font_path: Path | None,
) -> tuple[ImageFont.ImageFont, list[str]]:
    image = Image.new("RGB", (image_size, image_size))
    draw = ImageDraw.Draw(image)
    padding = int(image_size * DEFAULT_PADDING_RATIO)
    max_width = image_size - (padding * 2)
    max_height = image_size - (padding * 2)

    start_font_size = max(MIN_FONT_SIZE, min(DEFAULT_FONT_SIZE, image_size // 4))

    for font_size in range(start_font_size, MIN_FONT_SIZE - 1, -2):
        font = get_font(font_path, font_size)

        single_line = text.replace("\n", " ")
        single_width, single_height = measure_text_block(draw, [single_line], font)
        if single_width <= max_width and single_height <= max_height:
            return font, [single_line]

        wrapped_lines = wrap_text(draw, text, font, max_width)
        wrapped_width, wrapped_height = measure_text_block(draw, wrapped_lines, font)
        if wrapped_width <= max_width and wrapped_height <= max_height:
            return font, wrapped_lines

    font = get_font(font_path, MIN_FONT_SIZE)
    return font, wrap_text(draw, text, font, max_width)


def draw_centered_text(
    text: str,
    output_path: Path,
    image_size: int,
    background: str,
    text_color: str,
    font_path: Path | None,
) -> None:
    image = Image.new("RGB", (image_size, image_size), background)
    draw = ImageDraw.Draw(image)
    font, lines = fit_text_to_box(image_size, text, font_path)

    widths: list[int] = []
    boxes: list[tuple[int, int, int, int]] = []
    for line in lines:
        box = draw.textbbox((0, 0), line or " ", font=font)
        boxes.append(box)
        widths.append(box[2] - box[0])

    line_heights = [box[3] - box[1] for box in boxes]
    line_height = max(line_heights) if line_heights else 0
    line_spacing = max(4, int(line_height * LINE_SPACING_RATIO))
    block_height = sum(line_heights) + max(0, len(lines) - 1) * line_spacing

    y = (image_size - block_height) / 2
    for index, line in enumerate(lines):
        width = widths[index]
        x = (image_size - width) / 2
        draw.text((x, y), line, font=font, fill=text_color)
        y += line_heights[index] + line_spacing

    image.save(output_path, format="PNG")


def main() -> int:
    args = parse_args()
    xlsx_file = args.xlsx_file.expanduser().resolve()

    if not xlsx_file.exists():
        print(f"Input file not found: {xlsx_file}", file=sys.stderr)
        return 1

    if xlsx_file.suffix.lower() != ".xlsx":
        print("Input file must be an .xlsx file.", file=sys.stderr)
        return 1

    output_dir = args.output.expanduser().resolve() if args.output else xlsx_file.with_name(f"{xlsx_file.stem}_images")
    output_dir.mkdir(parents=True, exist_ok=True)

    values = list(iter_first_column_values(xlsx_file))
    if not values:
        print("No non-empty values found in the first column.", file=sys.stderr)
        return 1

    digits = max(3, int(math.log10(len(values))) + 1)

    for index, text in enumerate(values, start=1):
        filename = f"{index:0{digits}d}.png"
        output_path = output_dir / filename
        draw_centered_text(
            text=text,
            output_path=output_path,
            image_size=args.size,
            background=args.background,
            text_color=args.text_color,
            font_path=args.font,
        )

    print(f"Created {len(values)} image(s) in {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
